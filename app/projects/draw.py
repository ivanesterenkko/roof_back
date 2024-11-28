import base64
import matplotlib.pyplot as plt
from matplotlib import patches
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def draw_plan(lines, sheets, width):
    # Предварительное вычисление максимальных значений координат
    x_values = [sheet.x_start + width for sheet in sheets] + [line.x_end for line in lines]
    y_values = [sheet.y_start + sheet.length for sheet in sheets] + [line.y_end for line in lines]
    x_max = max(x_values)
    y_max = max(y_values)
    
    fig, ax = plt.subplots()

    # Общие параметры для стилей текста и рамок
    bbox_props = dict(facecolor='white', edgecolor='none', boxstyle='round,pad=0.3')
    text_props = dict(ha='center', va='center', fontsize=6, color='black', bbox=bbox_props)

    # Рисуем листы в виде прямоугольников
    for sheet in sheets:
        rectangle = patches.Rectangle(
            (sheet.x_start, sheet.y_start), 
            width,                      
            sheet.length,                  
            linewidth=1,                 
            edgecolor='gray',                  
            facecolor='none'                    
        )
        ax.add_patch(rectangle)
        ax.text(
            sheet.x_start + width / 2, 
            sheet.y_start + sheet.length / 2,  
            f"{sheet.length:.2f}", 
            **text_props
        )

    # Рисуем линии и их метки
    for line in lines:
        x1, y1, x2, y2 = line.x_start, line.y_start, line.x_end, line.y_end
        ax.plot([x1, x2], [y1, y2], color='black', linewidth=3)
        mid_x = (x1 + x2) / 2
        mid_y = (y1 + y2) / 2
        ax.text(
            mid_x, mid_y,
            line.name,
            fontweight='semibold',
            **text_props
        )
    
    # Устанавливаем пределы осей
    ax.set_xlim(0, x_max + 1)
    ax.set_ylim(0, y_max + 1)
    
    # Настраиваем сетку и пропорции
    ax.set_xticks(range(0, int(x_max) + 2))
    ax.set_yticks(range(0, int(y_max) + 2))
    ax.grid(visible=True, color='grey', linestyle='--', linewidth=0.5)
    ax.set_aspect('equal')
    
    # Сохраняем изображение в буфер в формате PNG
    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=300, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    image_bytes = buf.getvalue()
    
    image_base64 = base64.b64encode(image_bytes).decode('utf-8')

    return image_base64


def create_excel(data_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Спецификация"
    data = data_dict.dict()

    # Настройки шрифта и стиля
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    # # Ввод основной информации
    # ws.merge_cells("A1:F1")
    # ws["A1"] = f"Проект: {data['project_name']}, Адрес: {data['project_address']}"
    # ws["A1"].alignment = center_alignment
    # ws["A1"].font = Font(bold=True, size=14)

    # # Добавление метаинформации о кровле
    # roof_info = [
    #     ["Тип покрытия", data["roof_base"]["roof_type"]],
    #     ["Полезная ширина листа", data["roof_base"]["roof_useful_width"]],
    #     ["Полная ширина листа", data["roof_base"]["roof_overall_width"]],
    #     ["Длина волны", data["roof_base"]["roof_overlap"]],
    #     ["Макс. реальная длина листа", data["roof_base"]["roof_max_length"]],
    # ]

    # ws.append([])
    # ws.append(["Метаинформация"])
    # for key, value in roof_info:
    #     ws.append([key, value])

    # # Листы
    # ws.append([])
    # ws.append(["Длина листа (м)", "Количество (шт)"])
    # for length, count in data["sheets_amount"].items():
    #     ws.append([length, count])

    # # Скатные крыши
    # ws.append([])
    # ws.append(["Скат", "Общая площадь (м2)", "Полезная площадь (м2)"])
    # for slope in data["slopes"]:
    #     ws.append(
    #         [slope["slope_name"], slope["area_overall"], slope["area_usefull"]]
    #     )

    # # Доборные элементы
    # ws.append([])
    # ws.append(["Название", "Общая длина", "Количество", "Цена"])
    # for accessory in data["accessories"]:
    #     ws.append(
    #         [
    #             accessory["name"],
    #             accessory["overall_length"],
    #             accessory["amount"],
    #             accessory["price"],
    #         ]
    #     )

    # # Софиты и J-профили
    # ws.append([])
    # ws.append(["Название", "Общая длина", "Ширина", "Количество", "Цена"])
    # for soffit in data["sofits"]:
    #     ws.append(
    #         [soffit["name"], soffit["overall_length"], soffit["width"], soffit["amount"], soffit["price"]]
    #     )

    # # Саморезы
    # ws.append([])
    # ws.append(["Название", "Количество", "Цена"])
    # for screw in data["screws"]:
    #     ws.append([screw["name"], screw["amount"], screw["price"]])

    # # Применение стилей к заголовкам
    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.border = border
    #         if cell.row == 1:
    #             cell.font = header_font
    #             cell.alignment = center_alignment

    # Сохранение в байтовый поток
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output